from fastapi import FastAPI, APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import certifi
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone
import qrcode
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from fastapi import Query
from typing import Optional
from typing import Literal

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
# mongo_url = os.environ['MONGO_URL']
# client = AsyncIOMotorClient(mongo_url)
# db = client[os.environ['DB_NAME']]

# MongoDB connection (TLS safe)
mongo_url = os.environ["MONGO_URL"]
db_name = os.environ.get("DB_NAME", "rupya")

client = AsyncIOMotorClient(
    mongo_url,
    tls=True,
    tlsCAFile=certifi.where(),
    serverSelectionTimeoutMS=30000
)



db = client[db_name]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# ==================== MODELS ====================

class AgentCreate(BaseModel):
    name: str
    phone: str

class Agent(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    agentId: str
    name: str
    phone: str
    qrUrl: str
    qrImage: str  # Base64 encoded QR image
    isActive: bool = True
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class LoanApplicationCreate(BaseModel):
    agentId: str
    loanType: str
    fullName: str
    pan: str
    aadhaar: str
    mobile: str
    dob: str
    address: Optional[str] = ""
    employmentType: str
    monthlyIncome: float
    email: EmailStr
    status: Literal["logged", "approved", "rejected"] = "logged"     
    disbursedAmount: Optional[float] = None

class LoanApplication(BaseModel):
    model_config = ConfigDict(extra="ignore")

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    agentId: str
    loanType: str
    fullName: str
    pan: str
    aadhaar: str
    mobile: str
    dob: str
    address: Optional[str] = ""
    employmentType: str
    monthlyIncome: float
    email: EmailStr

    type: str = "loan"
    status: Literal["logged", "approved", "rejected"] = "logged"          
    disbursedAmount: Optional[float] = None
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


    
class InsuranceApplicationCreate(BaseModel):
    agentId: str
    fullName: str
    mobile: str
    insuranceType: str
    email: EmailStr
    
    status: Literal["logged", "approved", "rejected"] = "logged"
    
class InsuranceApplication(BaseModel):
    id: str
    agentId: str
    fullName: str
    mobile: str
    insuranceType: str
    email: EmailStr

    type: str = "insurance"
    status: Literal["logged", "approved", "rejected"] = "logged"      
    disbursedAmount: Optional[float] = None
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class AgentLoginRequest(BaseModel):
    agentId: str



# ==================== HELPER FUNCTIONS ====================

def generate_qr_code(url: str) -> str:
    """Generate QR code and return as base64 string"""
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(url)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Convert to base64
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    img_base64 = base64.b64encode(buffer.getvalue()).decode()
    
    return f"data:image/png;base64,{img_base64}"

# ==================== AGENT ROUTES ====================




@api_router.post("/agents/login")
async def agent_login(payload: AgentLoginRequest):
    agent = await db.agents.find_one(
        {"agentId": payload.agentId, "isActive": True},
        {"_id": 0}
    )

    if not agent:
        raise HTTPException(
            status_code=403,
            detail="Invalid or deactivated Agent ID"
        )

    return {
        "agentId": agent["agentId"],
        "name": agent["name"]
    }


@api_router.get("/insurance-applications", response_model=List[InsuranceApplication])
async def get_insurance_applications():
    raw = await db.insurance_applications.find({}, {"_id": 0}).to_list(10000)
    cleaned = []

    for d in raw:
        if isinstance(d.get("createdAt"), str):
            d["createdAt"] = datetime.fromisoformat(d["createdAt"])

        # 🔥 HARD NORMALIZATION
        if not d.get("status"):
            d["status"] = "logged"

        if d.get("disbursedAmount") is None:
            d["disbursedAmount"] = None

        cleaned.append(d)

    return cleaned



@api_router.patch("/agents/{agent_id}/deactivate")
async def deactivate_agent(agent_id: str):
    result = await db.agents.update_one(
        {"agentId": agent_id},
        {"$set": {"isActive": False}}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Agent not found")

    return {"message": "Agent deactivated successfully"}


@api_router.get("/applications/agent/{agent_id}")
async def get_all_agent_applications(agent_id: str):
    loans = await db.loan_applications.find(
        {"agentId": agent_id},
        {"_id": 0}
    ).to_list(10000)

    insurance = await db.insurance_applications.find(
        {"agentId": agent_id},
        {"_id": 0}
    ).to_list(10000)

    for item in loans + insurance:
        if isinstance(item.get("createdAt"), str):
            item["createdAt"] = datetime.fromisoformat(item["createdAt"])

    return loans + insurance

@api_router.patch("/applications/{app_id}")
async def update_application(app_id: str, payload: dict):
    updated = await db.loan_applications.update_one(
        {"id": app_id},
        {"$set": payload}
    )

    if updated.matched_count == 0:
        updated = await db.insurance_applications.update_one(
            {"id": app_id},
            {"$set": payload}
        )

    if updated.matched_count == 0:
        raise HTTPException(status_code=404, detail="Application not found")

    return {"message": "Application updated"}



@api_router.post("/agents", response_model=Agent)
async def create_agent(agent_input: AgentCreate):
    """Create a new agent with unique QR code"""
    # Generate unique agent ID
    agent_id = f"VND{str(uuid.uuid4())[:8].upper()}"
    
    # Check if agent ID already exists (unlikely but safe)
    existing = await db.agents.find_one({"agentId": agent_id}, {"_id": 0})
    if existing:
        agent_id = f"VND{str(uuid.uuid4())[:8].upper()}"
    
    # Get frontend URL from environment
    frontend_url = os.environ.get('FRONTEND_URL', 'http://localhost:3000')
    # qr_url = f"{frontend_url}/form?agentId={agent_id}"
    qr_url = f"{frontend_url}/scan?agentId={agent_id}"

    
    # Generate QR code
    qr_image = generate_qr_code(qr_url)
    
    # Create agent object
    agent = Agent(
        agentId=agent_id,
        name=agent_input.name,
        phone=agent_input.phone,
        qrUrl=qr_url,
        qrImage=qr_image
    )
    
    # Save to database
    doc = agent.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    await db.agents.insert_one(doc)
    
    return agent

@api_router.get("/agents", response_model=List[Agent])
async def get_agents():
    """Get all agents"""
    agents = await db.agents.find({}, {"_id": 0}).to_list(1000)
    
    # Convert ISO string timestamps back to datetime
    for agent in agents:
        if isinstance(agent['createdAt'], str):
            agent['createdAt'] = datetime.fromisoformat(agent['createdAt'])
    
    return agents

@api_router.get("/agents/{agent_id}", response_model=Agent)
async def get_agent(agent_id: str):
    """Get agent by ID"""
    agent = await db.agents.find_one({"agentId": agent_id}, {"_id": 0})
    
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
    
    if isinstance(agent['createdAt'], str):
        agent['createdAt'] = datetime.fromisoformat(agent['createdAt'])
    
    return agent

# ==================== LOAN APPLICATION ROUTES ====================

@api_router.post("/loan-applications", response_model=LoanApplication)
async def create_loan_application(application: LoanApplicationCreate):
    """Submit a new loan application"""
    # Verify agent exists
    agent = await db.agents.find_one(
        {"agentId": application.agentId, "isActive": True},
        {"_id": 0}
    )

    if not agent:
        raise HTTPException(
            status_code=403,
            detail="Agent is deactivated or invalid"
        )

    
    # Create application object
    app_obj = LoanApplication(**application.model_dump())
    
    # Save to database
    doc = app_obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    await db.loan_applications.insert_one(doc)
    
    return app_obj

@api_router.get("/loan-applications", response_model=List[LoanApplication])
async def get_loan_applications():
    raw = await db.loan_applications.find({}, {"_id": 0}).to_list(10000)
    cleaned = []

    for d in raw:
        if isinstance(d.get("createdAt"), str):
            d["createdAt"] = datetime.fromisoformat(d["createdAt"])

        # 🔥 HARD NORMALIZATION
        if not d.get("status"):
            d["status"] = "logged"

        if d.get("disbursedAmount") is None:
            d["disbursedAmount"] = None

        cleaned.append(d)

    return cleaned


    # """Get all loan applications with optional filters"""
    # query = {}
    
    # if agentId:
    #     query["agentId"] = agentId
    
    # if startDate or endDate:
    #     date_query = {}
    #     if startDate:
    #         date_query["$gte"] = startDate
    #     if endDate:
    #         date_query["$lte"] = endDate
    #     query["createdAt"] = date_query
    
    # applications = await db.loan_applications.find(query, {"_id": 0}).to_list(10000)
    
    # # Convert ISO string timestamps back to datetime
    # for app in applications:
    #     if isinstance(app['createdAt'], str):
    #         app['createdAt'] = datetime.fromisoformat(app['createdAt'])
    
    # return applications

@api_router.get("/loan-applications/agent/{agent_id}", response_model=List[LoanApplication])
async def get_agent_applications(agent_id: str):
    """Get all applications for a specific agent"""
    applications = await db.loan_applications.find(
        {"agentId": agent_id}, 
        {"_id": 0}
    ).to_list(10000)
    
    # Convert ISO string timestamps back to datetime
    for app in applications:
        if isinstance(app['createdAt'], str):
            app['createdAt'] = datetime.fromisoformat(app['createdAt'])
    
    return applications


@app.on_event("startup")
async def startup_db_check():
    await client.admin.command("ping")
    print("✅ MongoDB connected successfully")



@api_router.get("/loan-applications/export")
async def export_loan_applications(
    agentId: Optional[str] = Query(None)
):
    query = {}
    if agentId:
        query["agentId"] = agentId

    applications = await db.loan_applications.find(query, {"_id": 0}).to_list(10000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Loan Applications"

    headers = [
    "Application ID",     # column 1 -> id
    "Agent ID",           # column 2 -> agentId
    "Loan Type",          # column 3 -> loanType
    "Full Name",          # column 4 -> fullName
    "PAN",                # column 5 -> pan
    "Aadhaar",            # column 6 -> aadhaar
    "Mobile",             # column 7 -> mobile
    "Date of Birth",      # column 8 -> dob
    "Employment Type",    # column 9 -> employmentType
    "Monthly Income",     # column 10 -> monthlyIncome
    "Email",              # column 11 -> email
    "Status",             # column 12 -> status
    "Disbursed Amount",   # column 13 -> disbursedAmount
    "Submitted At"        # column 14 -> createdAt
    ]


    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)

    for row_idx, app in enumerate(applications, start=2):
        ws.cell(row=row_idx, column=1, value=app.get("id"))
        ws.cell(row=row_idx, column=2, value=app.get("agentId"))
        ws.cell(row=row_idx, column=3, value=app.get("loanType"))
        ws.cell(row=row_idx, column=4, value=app.get("fullName"))
        ws.cell(row=row_idx, column=5, value=app.get("pan"))
        ws.cell(row=row_idx, column=6, value=app.get("aadhaar"))
        ws.cell(row=row_idx, column=7, value=app.get("mobile"))
        ws.cell(row=row_idx, column=8, value=app.get("dob"))
        ws.cell(row=row_idx, column=9, value=app.get("employmentType"))
        ws.cell(row=row_idx, column=10, value=app.get("monthlyIncome"))
        ws.cell(row=row_idx, column=11, value=app.get("email"))
        ws.cell(row=row_idx, column=12, value=app.get("status"))
        ws.cell(row=row_idx, column=13, value=app.get("disbursedAmount") or "")
        ws.cell(row=row_idx, column=14, value=app.get("createdAt"))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=loan_applications.xlsx"}
    )
    
    
@api_router.post("/insurance-applications", response_model=InsuranceApplication)
async def create_insurance_application(app: InsuranceApplicationCreate):
    agent = await db.agents.find_one(
        {"agentId": app.agentId, "isActive": True}
    )

    if not agent:
        raise HTTPException(
            status_code=403,
            detail="Agent is deactivated or invalid"
        )

    ins_obj = InsuranceApplication(
        id=str(uuid.uuid4()),
        agentId=app.agentId,
        fullName=app.fullName,
        mobile=app.mobile,
        insuranceType=app.insuranceType,
        email=app.email,
        status="logged",
        disbursedAmount=None,
        createdAt=datetime.now(timezone.utc)
    )

    doc = ins_obj.model_dump()
    doc["createdAt"] = doc["createdAt"].isoformat()
    await db.insurance_applications.insert_one(doc)

    return ins_obj



@api_router.get("/insurance-applications/export")
async def export_insurance_applications():
    applications = await db.insurance_applications.find({}, {"_id": 0}).to_list(10000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Insurance Applications"

    headers = [
    "Agent ID",
    "Full Name",
    "Mobile",
    "Insurance Type",
    "Status",
    "Disbursed Amount",
    "Submitted At"
    ]


    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)

    for row_idx, app in enumerate(applications, start=2):
        ws.cell(row=row_idx, column=1, value=app.get("agentId"))
        ws.cell(row=row_idx, column=2, value=app.get("fullName"))
        ws.cell(row=row_idx, column=3, value=app.get("mobile"))
        ws.cell(row=row_idx, column=4, value=app.get("insuranceType"))
        ws.cell(row=row_idx, column=5, value=app.get("status"))
        ws.cell(row=row_idx, column=6, value=app.get("disbursedAmount") or "")
        ws.cell(row=row_idx, column=7, value=app.get("createdAt"))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=insurance.xlsx"}
    )

# ==================== BASIC ROUTES ====================

@api_router.get("/")
async def root():
    return {"message": "Loan QR System API"}

@api_router.get("/health")
async def health_check():
    return {"status": "healthy"}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
